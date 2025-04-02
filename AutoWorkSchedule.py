import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from datetime import datetime, timedelta
from itertools import cycle
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


class ShiftScheduler:
    def __init__(self, master):
        self.master = master
        master.title("智能排班系统 v3.4")
        self.employees = []
        self.shift_types = ['早班', '晚班', '休息']
        self.create_widgets()

        try:
            master.iconbitmap("resources/Gartoon.ico")
        except Exception as e:
            try:
                icon_img = tk.PhotoImage(file="resources/Gartoon.ico")
                master.wm_iconphoto(False, icon_img)
            except Exception as e:
                print(f"图标加载失败：{str(e)}")

    def create_widgets(self):
        # 输入区域
        input_frame = ttk.Frame(self.master, padding=10)
        input_frame.grid(row=0, column=0, sticky='w')

        # 员工信息输入
        ttk.Label(input_frame, text="姓  名:").grid(row=0, column=0)
        self.name_entry = ttk.Entry(input_frame, width=12)
        self.name_entry.grid(row=0, column=1, padx=2)

        ttk.Label(input_frame, text="手机号:").grid(row=0, column=2)
        self.phone_entry = ttk.Entry(input_frame, width=15)
        self.phone_entry.grid(row=0, column=3, padx=2)

        ttk.Label(input_frame, text="微  信:").grid(row=0, column=4)
        self.wechat_entry = ttk.Entry(input_frame, width=15)
        self.wechat_entry.grid(row=0, column=5, padx=2)

        ttk.Button(input_frame, text="添加员工", command=self.add_employee).grid(row=0, column=6, padx=5)
        ttk.Button(input_frame, text="删除选中", command=self.delete_employee).grid(row=0, column=7, padx=5)

        # 员工列表
        self.tree = ttk.Treeview(self.master, columns=('name', 'phone', 'wechat'),
                                 show='headings', height=10)
        self.tree.heading('name', text='姓名')
        self.tree.heading('phone', text='手机号')
        self.tree.heading('wechat', text='微信号')
        self.tree.column('name', width=100)
        self.tree.column('phone', width=120)
        self.tree.column('wechat', width=120)
        self.tree.grid(row=1, column=0, padx=10, pady=5, sticky='nsew')

        # 控制区域
        control_frame = ttk.Frame(self.master, padding=10)
        control_frame.grid(row=2, column=0, sticky='ew')

        # 日期范围输入
        ttk.Label(control_frame, text="开始日期:").grid(row=0, column=0)
        self.start_date_entry = ttk.Entry(control_frame, width=12)
        self.start_date_entry.grid(row=0, column=1, padx=2)
        self.start_date_entry.insert(0, "YYYY-MM-DD")

        ttk.Label(control_frame, text="结束日期:").grid(row=0, column=2)
        self.end_date_entry = ttk.Entry(control_frame, width=12)
        self.end_date_entry.grid(row=0, column=3, padx=2)
        self.end_date_entry.insert(0, "YYYY-MM-DD")

        ttk.Button(control_frame, text="生成排班", command=self.generate_schedule).grid(row=0, column=4, padx=5)
        ttk.Button(control_frame, text="导出Excel", command=self.export_excel).grid(row=0, column=5, padx=5)

    def add_employee(self):
        """添加员工到列表"""
        name = self.name_entry.get().strip()
        phone = self.phone_entry.get().strip()
        wechat = self.wechat_entry.get().strip()

        if not name:
            messagebox.showwarning("输入错误", "请输入员工姓名")
            return

        if not phone.isdigit() or len(phone) != 11:
            messagebox.showwarning("输入错误", "请输入有效的11位手机号")
            return

        self.employees.append({'name': name, 'phone': phone, 'wechat': wechat or "无"})
        self.tree.insert('', 'end', values=(name, phone, wechat))
        self.clear_entries()

    def delete_employee(self):
        """删除选中员工"""
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("操作提示", "请先选择要删除的员工")
            return

        for item in selected_item:
            # 从数据源删除
            item_values = self.tree.item(item)['values']
            self.employees = [emp for emp in self.employees
                              if emp['name'] != item_values or
                              emp['phone'] != item_values]
            # 从界面删除
            self.tree.delete(item)

    def clear_entries(self):
        """清空输入框"""
        self.name_entry.delete(0, 'end')
        self.phone_entry.delete(0, 'end')
        self.wechat_entry.delete(0, 'end')

    def generate_schedule(self):
        """生成排班表"""
        try:
            start_date = datetime.strptime(self.start_date_entry.get(), "%Y-%m-%d")
            end_date = datetime.strptime(self.end_date_entry.get(), "%Y-%m-%d")
            if start_date > end_date:
                raise ValueError("开始日期不能晚于结束日期")
        except Exception as e:
            messagebox.showerror("输入错误", f"日期错误：{str(e)}\n请使用YYYY-MM-DD格式")
            return

        if not self.employees:
            messagebox.showerror("错误", "请先添加员工")
            return

        # 生成日期序列
        self.date_sequence = []
        current_date = start_date
        while current_date <= end_date:
            self.date_sequence.append(current_date)
            current_date += timedelta(days=1)

        # 生成排班数据
        schedule_data = []
        shift_cycle = cycle(self.shift_types)

        for emp in self.employees:
            shifts = {}
            current_shift = next(shift_cycle)
            for idx, date in enumerate(self.date_sequence):
                if idx % len(self.shift_types) == 0:
                    current_shift = next(shift_cycle)
                shifts[date.strftime("%Y-%m-%d")] = current_shift
                current_shift = next(shift_cycle)

            schedule_data.append({
                '员工姓名': emp['name'],
                '手机号': emp['phone'],
                '微信号': emp['wechat'],
                **shifts
            })

        self.df = pd.DataFrame(schedule_data)
        messagebox.showinfo("成功", "排班表已生成，请导出Excel文件")

    def export_excel(self):
        """导出到Excel文件"""
        if not hasattr(self, 'df'):
            messagebox.showerror("错误", "请先生成排班表")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[("Excel文件", "*.xlsx"), ("All Files", "*.*")]
        )

        if not file_path:
            return

        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # 排班明细表
                self.df.to_excel(writer, index=False, sheet_name='排班明细')
                self.format_detail_sheet(writer.sheets['排班明细'])

                # 日历视图
                self.create_calendar_view(writer)

            messagebox.showinfo("成功", f"文件已保存至：{file_path}")
        except Exception as e:
            messagebox.showerror("保存失败", f"导出失败：{str(e)}")

    def format_detail_sheet(self, ws):
        """格式化明细表"""
        for col_idx, col in enumerate(ws.columns, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 14
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    def create_calendar_view(self, writer):
        """创建日历视图"""
        calendar_data = []
        date_columns = [d.strftime("%Y-%m-%d") for d in self.date_sequence]

        for date_obj in self.date_sequence:
            date_str = date_obj.strftime("%Y-%m-%d")
            day_record = {'日期': date_str}

            # 遍历所有员工获取当天班次
            for employee in self.employees:
                # 精确匹配员工数据
                emp_record = self.df[
                    (self.df['员工姓名'] == employee['name']) &
                    (self.df['手机号'] == employee['phone'])
                    ]

                if not emp_record.empty:
                    try:
                        # 正确获取日期列的值
                        shift = emp_record[date_str].values
                    except KeyError:
                        shift = "无排班"
                else:
                    shift = "员工不存在"

                day_record[employee['name']] = shift

            calendar_data.append(day_record)

        calendar_df = pd.DataFrame(calendar_data)
        calendar_df.to_excel(writer, index=False, sheet_name='日历视图')
        self.format_calendar_sheet(writer.sheets['日历视图'])

    def format_calendar_sheet(self, ws):
        """格式化日历表"""
        for col_idx, col in enumerate(ws.columns, 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max(max_len + 2, 10)
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.freeze_panes = 'B2'


if __name__ == "__main__":
    root = tk.Tk()
    app = ShiftScheduler(root)
    root.mainloop()
